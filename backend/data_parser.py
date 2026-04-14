"""
Data parser for inventory CSV/Excel files
Uses only stdlib csv module (no pandas dependency)
"""

import os
import csv
from typing import Dict, List, Any
from datetime import datetime


def parse_inventory_file(file_path: str, skip_header_rows: int = 0) -> Dict[str, Any]:
    """
    Парсинг файла инвентаризации (CSV или Excel)

    Args:
        file_path: Путь к файлу
        skip_header_rows: Количество строк для пропуска в начале (для SAP-отчётов)

    Возвращает структуру:
    {
        "parsed_at": "2026-04-11T...",
        "file_name": "...",
        "groups": [...],
        "summary": {...}
    }
    """
    file_name = os.path.basename(file_path)
    ext = os.path.splitext(file_name)[1].lower()

    rows = []
    if ext == '.csv':
        rows = _read_csv_file(file_path)
    elif ext in ['.xlsx', '.xls']:
        rows = _read_excel_file(file_path, skip_header_rows=skip_header_rows)
    else:
        raise ValueError(f"Неподдерживаемый формат: {ext}")
    
    # Обрабатываем строки
    groups = []
    for row in rows:
        # Пропускаем строки без Группа ID
        group_id = str(row.get('Группа ID', '')).strip()
        if not group_id or group_id == 'nan' or group_id == '':
            continue
        
        # Парсим долю
        доля_raw = row.get('Доля', '')

        # Excel может хранить проценты как десятичную дробь (0.5 вместо 50)
        if isinstance(доля_raw, (int, float)):
            доля = float(доля_raw)
            if 0 < доля <= 1:
                доля = доля * 100  # Конвертируем 0.5 → 50
            доля = round(доля, 1)
        else:
            доля_str = str(доля_raw).strip().replace('%', '')
            try:
                доля = float(доля_str) if доля_str and доля_str != 'nan' else 0.0
            except ValueError:
                доля = 0.0
        
        def safe_int(val, default=0):
            try:
                v = str(val).strip()
                return int(float(v)) if v and v != 'nan' else default
            except (ValueError, TypeError):
                return default
        
        group = {
            'Дивизион ASM': str(row.get('Дивизион ASM', '')).strip(),
            'Город': str(row.get('Город', '')).strip(),
            'Магазин': str(row.get('Магазин', '')).strip(),
            'Дивизион SAP': str(row.get('Дивизион SAP', '')).strip(),
            'Категория': str(row.get('Категория', '')).strip(),
            'Подкатегория': str(row.get('Подкатегория', '')).strip(),
            'Планнейм': str(row.get('Планнейм', '')).strip(),
            'Группа ID': group_id,
            'Группа': str(row.get('Группа', '')).strip(),
            'Дата пересчета': str(row.get('Дата пересчета', '')).strip(),
            'Частота подсчета': str(row.get('Частота подсчета', '')).strip(),
            'Товаров в группе': safe_int(row.get('Товаров в группе')),
            'Штук в группе': safe_int(row.get('Штук в группе')),
            'Найдено': safe_int(row.get('Найдено')),
            'Излишки': safe_int(row.get('Излишки')),
            'Недостачи': safe_int(row.get('Недостачи')),
            'Брак': safe_int(row.get('Брак')),
            'Посчитано вручную': safe_int(row.get('Посчитано вручную')),
            'Посчитано групп': safe_int(row.get('Посчитано групп')),
            'Всего групп': safe_int(row.get('Всего групп')),
            'Доля': доля,
        }
        
        groups.append(group)
    
    # Считаем summary
    total = len(groups)
    counted = sum(1 for g in groups if g['Доля'] == 100)
    partial = sum(1 for g in groups if 0 < g['Доля'] < 100)
    not_counted = sum(1 for g in groups if g['Доля'] == 0)
    
    result = {
        'parsed_at': datetime.now().isoformat(),
        'file_name': file_name,
        'groups': groups,
        'summary': {
            'total_groups': total,
            'counted_groups': counted,
            'partial_groups': partial,
            'not_counted_groups': not_counted,
            'counted_percentage': round((counted / total * 100), 1) if total > 0 else 0,
        }
    }
    
    return result


def _read_csv_file(file_path: str) -> List[Dict[str, str]]:
    """
    Чтение CSV файла с автоопределением разделителя
    """
    separators = [';', ',', '\t', '|']
    
    for sep in separators:
        try:
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                first_line = f.readline().strip()
            
            count = first_line.count(sep)
            if count > 0:
                with open(file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.DictReader(f, delimiter=sep)
                    rows = list(reader)
                    if len(rows) > 0 and len(rows[0]) > 1:
                        return _normalize_columns(rows)
        except Exception:
            continue
    
    # Fallback
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        return _normalize_columns(list(reader))


def _read_excel_file(file_path: str, skip_header_rows: int = 0) -> List[Dict[str, str]]:
    """
    Чтение Excel файла через openpyxl
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active

    rows = []
    headers = None
    row_num = 0

    for row in ws.iter_rows(values_only=True):
        row_num += 1
        # Пропускаем первые N строк (служебные строки SAP: дата, имя отчёта и т.д.)
        if skip_header_rows > 0 and row_num <= skip_header_rows:
            continue

        # Первая строка после пропуска — это заголовки
        if headers is None:
            headers = [str(h).strip() if h else '' for h in row]
        else:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    row_dict[headers[i]] = val if val is not None else ''
            rows.append(row_dict)

    wb.close()
    return _normalize_columns(rows)


def _normalize_columns(rows: List[Dict[str, str]]) -> List[Dict[str, str]]:
    """
    Нормализация имён колонок — маппинг различных вариантов на стандартные
    """
    column_mapping = {
        'Дивизион ASM': 'Дивизион ASM',
        'ASM': 'Дивизион ASM',
        'Город': 'Город',
        'Магазин': 'Магазин',
        'Дивизион SAP': 'Дивизион SAP',
        'Материал: Категория': 'Категория',
        'Категория': 'Категория',
        'Подкатегория': 'Подкатегория',
        'Планнейм': 'Планнейм',
        'Группа ID': 'Группа ID',
        'Группа': 'Группа',
        'Дата пересчета': 'Дата пересчета',
        'Частота подсчета': 'Частота подсчета',
        'Товаров в группе': 'Товаров в группе',
        'Штук в группе': 'Штук в группе',
        'Найдено': 'Найдено',
        'Излишки': 'Излишки',
        'Недостачи': 'Недостачи',
        'Брак': 'Брак',
        'Посчитано вручную': 'Посчитано вручную',
        'Посчитано групп': 'Посчитано групп',
        'Всего групп': 'Всего групп',
        'Доля посчитанных товарных групп': 'Доля',
    }
    
    normalized = []
    for row in rows:
        new_row = {}
        for key, value in row.items():
            normalized_key = column_mapping.get(key, key)
            new_row[normalized_key] = value
        normalized.append(new_row)
    
    return normalized
